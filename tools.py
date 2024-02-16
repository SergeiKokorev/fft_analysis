import os
import csv
import numpy as np

from openpyxl import load_workbook


def get_data(file: str):
    
    x = []
    y = []
    ext = os.path.splitext(os.path.split(file)[1])[1]

    with open(file, 'r', newline='') as f:

        if ext == '.csv':
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                try:
                    xi, yi = float(row[0]), float(row[1])
                    x.append(xi)
                    y.append(yi)
                except ValueError:
                    continue
                except AttributeError:
                    raise AttributeError('Unsupported value type')

        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(filename=file)
            sheet_name = wb.sheetnames[0]
            rows = list(wb[sheet_name].values)

            for row in rows:
                try:
                    xi = float(row[0])
                    yi = float(row[1])
                    x.append(xi)
                    y.append(yi)
                except ValueError:
                    continue
                except AttributeError:
                    raise AttributeError('unsupported value type')
    
    return np.array(x), np.array(y)
